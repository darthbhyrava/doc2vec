import openpyxl
import nltk
import os

path = os.getcwd()
files = os.listdir('/home/darthbhyrava/2018/doc2vec/input/')
files_xls = [f for f in files if f[-4:] == 'xlsx']
print files_xls

text = {}
party = {}

index = 1
for f in files_xls:
    print "FILE: {0}".format(f)
    wb = openpyxl.load_workbook(path+'/input/{0}'.format(f))
    print "done"
    ws = wb.active
    text_column = ws['F']
    party_column = ws['D']
    rows = len(text_column)
    for row in range(rows):
        print index
        text[index] = text_column[row].value
        party[index] = party_column[row].value
        index += 1
    wb.close()
print party